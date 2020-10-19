py begin
# encoding:utf-8
import pandas as pd
import sys
import codecs
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, colors
import csv
import os
#sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())
reload(sys)
sys.setdefaultencoding('utf8')  

def get_userinfo():
    birth_weight_file = '/root/record_scores/info.csv'  # 路径固定
    with open(birth_weight_file) as csvfile:
        mLines = csvfile.readlines()
    targetLine = mLines[-1]
    file_name = targetLine.split(',')[0]
    course_series = targetLine.split(',')[1]
    username = targetLine.split(',')[3]
    password = targetLine.split(',')[4]
    return file_name,course_series,username,password

file_name,course_series,username,password=get_userinfo()
create_csv='/root/record_scores/'+course_series+'.csv'
check_file='/root/rpa_file/'+course_series+'-check.xlsx'
check_png='/root/rpa_file/'+course_series+'.png'

#先判断是否有该课程序号的csv文件，xlsx文件及png文件。有则删除
if (os.path.exists(create_csv)):
    os.remove(create_csv)
if (os.path.exists(check_file)):
    os.remove(check_file)
if (os.path.exists(check_png)):
    os.remove(check_png)

#excel文件转csv格式,并统一转xlsx格式
def get_student_list(path):  # 读取excel
    df = pd.read_excel(path)
    datadict = df.fillna('').to_dict(orient='record')
    return datadict, df

student_list, excel_tmp = get_student_list(file_name)  # 上传的原文件
#excel_tmp.to_csv(create_csv,encoding='utf-8-sig')
excel_tmp.to_excel(check_file)
wb=load_workbook(check_file)
sheet_names = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheet_names[0])
ws.delete_cols(1)
wb.save(check_file)


wb = load_workbook(check_file)
ws = wb.worksheets[0]
rows=ws.max_row   #获取行数
cols=ws.max_column
for i in range(1,cols+1):
    name=ws.cell(row=1, column=i).value
    if name=='平时成绩Ⅰ':
        ws.cell(row=1, column=i, value='平时成绩1')
    if name=='平时成绩Ⅱ':
        ws.cell(row=1, column=i, value='平时成绩2')
    if name=='平时成绩Ⅲ':
        ws.cell(row=1, column=i, value='平时成绩3')
wb.save(check_file)

student_list, excel_tmp = get_student_list(check_file)
excel_tmp.to_csv(create_csv,encoding='utf-8-sig')

#data_xls = pd.read_excel(check_file, index_col=0)
#data_xls.to_csv(create_csv, encoding='utf-8')


py finish



 
  






