py begin
# encoding:utf-8
import pandas as pd
import sys
import codecs
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, colors
import csv
import os
#sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())
reload(sys)
sys.setdefaultencoding('utf8')  

def get_userinfo():
    birth_weight_file = '/root/record_scores/info.csv'  # 路径固定
    with open(birth_weight_file) as csvfile:
        mLines = csvfile.readlines()
    targetLine = mLines[-1]
    file_name = targetLine.split(',')[0]
    course_series = targetLine.split(',')[1]
    username = targetLine.split(',')[3]
    password = targetLine.split(',')[4]
    return file_name,course_series,username,password

file_name,course_series,username,password=get_userinfo()
create_csv='/root/record_scores/'+course_series+'.csv'
check_file='/root/rpa_file/'+course_series+'-check.xlsx'
check_png='/root/rpa_file/'+course_series+'.png'

# 获取period.csv写入列位置
birth_weight_file = '/root/record_scores/period.csv'  # 路径固定
with open(birth_weight_file) as csvfile:
    period = csvfile.readlines()
print(period)
period_1=[]
period_2=[]
for p in period:
    p=p.split('\r\n')
    p1=p[0]
    period_1.append(p1)
print(period_1)
for p in period_1:
    p=p.split('|')
    p1=p[1]
    period_2.append(p1)
print(period_2)
number_col=period_2[0]
print(number_col)


wb = load_workbook(check_file)
ws = wb.worksheets[0]
rows=ws.max_row   #获取行数
cols=ws.max_column

numbers=[]
for i in range(2,rows+1):
    for index in period_1[1:]:
        p=index.split('|')
        index=int(p[1])
        c_index=int(p[0])
        s=ws.cell(row=i, column=index).value
        if s==None:
            ws.cell(row=i, column=index, value=0)
        else:
            score = float(ws.cell(row=i, column=index).value)
            score = int(score+0.5)
            ws.cell(row=i, column=index, value=score)

wb.save(check_file)

#excel文件转csv格式,并统一转xlsx格式
def get_student_list(path):  # 读取excel
    df = pd.read_excel(path)
    datadict = df.fillna('').to_dict(orient='record')
    return datadict, df


student_list, excel_tmp = get_student_list(check_file)  # 上传的原文件
excel_tmp.to_csv(create_csv,encoding='utf-8-sig')

py finish



 
  






