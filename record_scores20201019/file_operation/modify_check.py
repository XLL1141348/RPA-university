py begin
# encoding:utf-8
import pandas as pd
import sys
import codecs
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, colors
import csv
import os
#sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())
reload(sys)
sys.setdefaultencoding('utf8') 


def get_userinfo():
    birth_weight_file = '/root/record_scores/info.csv'  # 路径固定
    with open(birth_weight_file) as csvfile:
        mLines = csvfile.readlines()
    targetLine = mLines[-1]
    file_name = targetLine.split(',')[0]
    course_series = targetLine.split(',')[1]
    username = targetLine.split(',')[3]
    password = targetLine.split(',')[4]
    return file_name,course_series,username,password


file_name,course_series,username,password=get_userinfo()
check_file='/root/rpa_file/'+course_series+'-check.xlsx'
fill_red = PatternFill("solid", fgColor="FF0000")
fill_green = PatternFill("solid", fgColor="C1F4B8")

# 获取period.csv写入列位置
birth_weight_file = '/root/record_scores/period.csv'  # 路径固定
with open(birth_weight_file) as csvfile:
    period = csvfile.readlines()
print(period)
period_1=[]
period_2=[]
for p in period:
    p=p.split('\r\n')
    p1=p[0]
    period_1.append(p1)
print(period_1)
for p in period_1:
    p=p.split('|')
    p1=p[1]
    period_2.append(p1)
print(period_2)
number_col=period_2[0]
print(number_col)









if (os.path.exists('/root/record_scores/record.csv')):  
    birth_weight_file = '/root/record_scores/record.csv'  # 路径固定
    with open(birth_weight_file) as csvfile:
        check = csvfile.readlines()

birth_weight_file = '/root/record_scores/web_number.csv'  # 路径固定
with open(birth_weight_file) as csvfile:
    web_number = csvfile.readlines()


wb = load_workbook(check_file)
ws = wb.worksheets[0]
rows=ws.max_row   #获取行数
cols=ws.max_column






# 根据check.csv文件校对成绩

for i in range(2,rows+1):
    number = ws.cell(row=i, column=int(number_col)).value
    for index in period_1[1:]:
        p=index.split('|')
        index=int(p[1])
        c_index=int(p[0])
        ws[i][index-1].fill = fill_green
wb.save(check_file)
# for c in check:
#     c=c.split(',')
#     c_number=c[0]
#     for i in range(2,rows+1):
#         number = ws.cell(row=i, column=int(number_col)).value
#         if int(number)==int(c_number):
#             for index in period_1[1:]:
#                 p=index.split('|')
#                 index=int(p[1])
#                 c_index=int(p[0])
#                 score=ws.cell(row=i, column=index).value
#                 if score is not None and int(c[c_index]) == int(score):
#                     ws[i][index-1].fill = fill_green
#                 elif score is not None and int(c[c_index]) != int(score):
#                     ws[i][index-1].fill = fill_red
# wb.save(check_file)

# 获取错误学号
if (os.path.exists('/root/record_scores/wrong.csv')):   
    birth_weight_file = '/root/record_scores/wrong.csv'  # 路径固定
    with open(birth_weight_file) as csvfile:
        wrong = csvfile.readlines()

    for w in wrong:
        for i in range(2,rows+1):
            number = ws.cell(row=i, column=int(number_col)).value
            if int(number)==int(w):
                for index in period_2[1:]:
                    index=int(index)
                    if  ws.cell(row=i, column=index).value is not None:
                        ws[i][index-1].fill = fill_red
    wb.save(check_file)


# 根据web_number.csv文件校对网页上是否有多余学号
numbers=[]
for i in range(2,rows+1):
    number = str(ws.cell(row=i, column=int(number_col)).value)
    ws.cell(row=i, column=int(number_col), value=number)
wb.save(check_file)
for i in range(2,rows+1):
    number = ws.cell(row=i, column=int(number_col)).value
    numbers.append(number)


web_number_1=[]
for w in web_number:
    w=w.split('\r\n')
    w1=w[0]
    if w1 not in numbers:
        web_number_1.append(w1)

if len(web_number_1)>0:
    wb.create_sheet(title=u'页面多出学号',index=1)
    ws_1 = wb.worksheets[1]
    for x in range(1,len(web_number_1)+1):
        ws_1.cell(row=x, column=1, value=web_number_1[x-1])
wb.save(check_file)


py finish



